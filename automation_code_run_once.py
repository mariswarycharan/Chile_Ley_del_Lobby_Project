from data_scraping import scrape_data
from storygeneration import generate_narratives
from faiss_db import create_vector_store
import datetime
import subprocess
import openpyxl
from openpyxl import Workbook

def commit_and_push_all_changes(commit_message):
    try:
        # Stage all changes
        subprocess.run(["git", "add", "."], check=True)

        # Check if there is anything to commit
        status_result = subprocess.run(
            ["git", "status", "--porcelain"],
            check=True,
            capture_output=True,
            text=True
        )

        if status_result.stdout.strip() == "":
            print("Nothing to commit. Working tree is clean.")
        else:
            # Commit with provided message
            subprocess.run(["git", "commit", "-m", commit_message], check=True)

        # Pull latest changes from the remote repository
        subprocess.run(["git", "pull"], check=True)

        # Push to current branch
        subprocess.run(["git", "push"], check=True)

        print("All changes committed and pushed successfully.")
    except subprocess.CalledProcessError as e:
        print(f"An error occurred: {e}")

def update_excel_with_last_updated(timestamp, excel_file="update_log.xlsx", sheet_name="Log"):
    try:
        wb = openpyxl.load_workbook(excel_file)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)

    # Write header and timestamp
    ws["A1"] = "Last Updated"
    ws["B1"] = timestamp.strftime("%Y-%m-%d %H:%M:%S")

    wb.save(excel_file)
    print(f"Updated Excel log: {excel_file} [{sheet_name}!B1] = {ws['B1'].value}")

# Function to execute the script
def run_script():
    print("Running scheduled script...")
    
    # Step 1 : Scrape data from the website
    print("Scraping data from the website...")
    #scrape_data()
    print("Scraping completed.")
    
    # Step 2 : Generate narratives from the scraped data
    print("Generating narratives from the scraped data...")
    #generate_narratives()
    print("Generating narratives completed.")
    
    # Step 3 : Create vector store
    print("Creating vector store...")
    #create_vector_store()
    print("Vector store creation completed.")
    
    # Step 4 : push chnages to repository
    print("Pushing changes to repository...")
    now = datetime.datetime.now()
    message = "data updated on " + str(datetime.datetime.now())
    commit_and_push_all_changes(message)
    print("Changes pushed to repository.")

    # Step 5: Log last-updated timestamp into Excel
    update_excel_with_last_updated(now)

    print("Script execution completed.")


if __name__ == "__main__":
    # Run the script
    run_script()